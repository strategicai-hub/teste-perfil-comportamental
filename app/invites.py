"""Leitura da planilha de colaboradores enviada pelo gestor.

Aceita `.xlsx` (openpyxl) e `.csv` (stdlib, separador `,` ou `;`). Espera as colunas
**Nome**, **E-mail** e **Área** — o cabeçalho é reconhecido sem depender de acento,
caixa ou pontuação, então `E-mail`, `email` e `EMAIL` valem a mesma coisa.

A leitura é dividida em três passos para permitir o wizard de mapeamento quando a
planilha do cliente não usa esses nomes:

1. `ler_grade()`   — devolve a planilha como matriz, sem interpretar nada;
2. `sugerir_mapa()`— tenta adivinhar qual coluna é qual pelos sinônimos conhecidos;
3. `mapear_e_validar()` — com o mapa (adivinhado ou escolhido na tela), valida linha
   a linha e devolve `(validas, erros)`.

`parse_planilha()` continua existindo como atalho dos três passos, com o mesmo
comportamento de antes: nunca levanta por linha ruim, devolve as válidas e os erros
para a tela mostrar a prévia antes de gravar qualquer coisa.
"""

import csv
import io
import re
import unicodedata

MAX_LINHAS = 2000
# Colunas extras a mais que isso não cabem no hidden input do wizard sem estourar o
# tamanho do form — a tela avisa para remover as sobras.
MAX_COLUNAS = 30

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[a-zA-Z]{2,}$")


def _normaliza_cabecalho(valor: object) -> str:
    """'E-mail ' → 'email'; 'Área' → 'area'. Tolera acento, caixa e pontuação."""
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", texto)


# Sinônimos aceitos para cada coluna, já normalizados.
_COLUNAS = {
    "nome": {"nome", "nomecompleto", "colaborador", "funcionario", "name"},
    "email": {"email", "emails", "correio", "correioeletronico", "mail"},
    "area": {"area", "setor", "departamento", "areasetor", "equipe", "time"},
}


def _mapear_colunas(cabecalho: list[object]) -> dict[str, int]:
    """Índice de cada coluna conhecida no cabeçalho da planilha."""
    mapa: dict[str, int] = {}
    for i, celula in enumerate(cabecalho):
        chave = _normaliza_cabecalho(celula)
        for campo, sinonimos in _COLUNAS.items():
            if chave in sinonimos and campo not in mapa:
                mapa[campo] = i
    return mapa


def _linhas_do_xlsx(conteudo: bytes) -> list[list[object]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    linhas = [list(row) for row in ws.iter_rows(max_row=MAX_LINHAS + 1, values_only=True)]
    wb.close()
    return linhas


def _linhas_do_csv(conteudo: bytes) -> list[list[object]]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            texto = conteudo.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Não consegui ler o arquivo. Salve como UTF-8 e tente de novo.")

    amostra = texto[:4096]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=",;\t")
        delimitador = dialect.delimiter
    except csv.Error:
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","

    leitor = csv.reader(io.StringIO(texto), delimiter=delimitador)
    return [linha for _, linha in zip(range(MAX_LINHAS + 1), leitor)]


ROTULOS = {"nome": "Nome", "email": "E-mail", "area": "Área"}


def ler_grade(nome_arquivo: str, conteudo: bytes) -> list[list[str]]:
    """Planilha como matriz de strings, sem interpretar cabeçalho.

    É o insumo do wizard de mapeamento: a tela precisa mostrar as colunas como elas
    vieram, mesmo quando os nomes não são reconhecidos.
    """
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".xlsx"):
        linhas = _linhas_do_xlsx(conteudo)
    elif nome.endswith((".csv", ".txt")):
        linhas = _linhas_do_csv(conteudo)
    else:
        raise ValueError("Formato não suportado. Envie um arquivo .xlsx ou .csv.")

    grade = [
        [str(c or "").strip() for c in linha]
        for linha in linhas
        if any(str(c or "").strip() for c in linha)
    ]
    if not grade:
        raise ValueError("A planilha está vazia.")
    if len(grade) - 1 > MAX_LINHAS:
        raise ValueError(f"A planilha tem mais de {MAX_LINHAS} linhas. Divida em arquivos menores.")
    largura = max(len(l) for l in grade)
    if largura > MAX_COLUNAS:
        raise ValueError(
            f"A planilha tem {largura} colunas — o limite é {MAX_COLUNAS}. "
            "Apague as colunas que não são Nome, E-mail e Área e envie de novo."
        )
    # Normaliza a largura: linha curta com célula faltando não pode virar IndexError.
    return [linha + [""] * (largura - len(linha)) for linha in grade]


def sugerir_mapa(grade: list[list[str]]) -> dict[str, int]:
    """Mapa campo→índice adivinhado pelo cabeçalho. Pode vir incompleto."""
    return _mapear_colunas(grade[0]) if grade else {}


def campos_faltando(mapa: dict[str, int]) -> list[str]:
    return [c for c in ("nome", "email", "area") if c not in mapa]


def colunas_para_escolha(grade: list[list[str]]) -> list[dict]:
    """Colunas com título e exemplos, para o gestor escolher qual é qual na tela."""
    if not grade:
        return []
    saida = []
    for i, titulo in enumerate(grade[0]):
        exemplos = [linha[i] for linha in grade[1:4] if i < len(linha) and linha[i]]
        saida.append({"indice": i, "titulo": titulo or f"Coluna {i + 1}", "exemplos": exemplos})
    return saida


def mapear_e_validar(grade: list[list[str]], mapa: dict[str, int]) -> tuple[list[dict], list[str]]:
    """Aplica o mapa na grade e devolve `(linhas_validas, erros)`.

    Cada linha válida é `{"nome", "email", "area"}`. E-mails repetidos dentro do
    próprio arquivo são reportados como erro e mantidos só na primeira ocorrência.
    """
    faltando = campos_faltando(mapa)
    if faltando:
        nomes = ", ".join(ROTULOS[c] for c in faltando)
        raise ValueError(f"Faltou indicar qual coluna é: {nomes}.")

    validas: list[dict] = []
    erros: list[str] = []
    vistos: set[str] = set()

    for numero, linha in enumerate(grade[1:], start=2):
        def celula(campo: str) -> str:
            i = mapa[campo]
            return linha[i].strip() if i < len(linha) else ""

        nome_col, email, area = celula("nome"), celula("email").lower(), celula("area")

        if not any((nome_col, email, area)):
            continue
        if not email:
            erros.append(f"Linha {numero}: sem e-mail.")
            continue
        if not _EMAIL_RE.match(email):
            erros.append(f"Linha {numero}: e-mail inválido ({email}).")
            continue
        if email in vistos:
            erros.append(f"Linha {numero}: e-mail repetido na planilha ({email}).")
            continue
        if not nome_col:
            erros.append(f"Linha {numero}: sem nome ({email}).")
            continue
        if not area:
            erros.append(f"Linha {numero}: sem área ({email}).")
            continue

        vistos.add(email)
        validas.append({"nome": nome_col, "email": email, "area": area})

    return validas, erros


def parse_colado(texto: str) -> tuple[list[dict], list[str]]:
    """Lê a lista colada no campo de texto: `Nome; email; Área` por linha.

    Aceita `;`, tabulação ou vírgula como separador — o gestor cola de onde tiver.
    """
    validas: list[dict] = []
    erros: list[str] = []
    vistos: set[str] = set()
    for numero, bruta in enumerate((texto or "").splitlines(), start=1):
        linha = bruta.strip()
        if not linha:
            continue
        for sep in (";", "\t", ","):
            if sep in linha:
                partes = [p.strip() for p in linha.split(sep)]
                break
        else:
            partes = [linha]
        if len(partes) < 3:
            erros.append(f"Linha {numero}: informe nome, e-mail e área separados por ponto e vírgula.")
            continue
        nome, email, area = partes[0], partes[1].lower(), partes[2]
        if not _EMAIL_RE.match(email):
            erros.append(f"Linha {numero}: e-mail inválido ({email}).")
            continue
        if email in vistos:
            erros.append(f"Linha {numero}: e-mail repetido ({email}).")
            continue
        if not nome or not area:
            erros.append(f"Linha {numero}: falta o nome ou a área ({email}).")
            continue
        vistos.add(email)
        validas.append({"nome": nome, "email": email, "area": area})
    return validas, erros


def modelo_xlsx() -> bytes:
    """Planilha modelo para download, já com o cabeçalho que o sistema reconhece."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"

    ws.append(["Nome", "E-mail", "Área"])
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="0F766E")
        celula.alignment = Alignment(horizontal="center")

    for linha in (
        ["Ana Souza", "ana@empresa.com.br", "Comercial"],
        ["Bruno Lima", "bruno@empresa.com.br", "Operações"],
        ["Carla Dias", "carla@empresa.com.br", "Administrativo"],
    ):
        ws.append(linha)

    for coluna, largura in zip("ABC", (32, 38, 24)):
        ws.column_dimensions[coluna].width = largura
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_planilha(nome_arquivo: str, conteudo: bytes) -> tuple[list[dict], list[str]]:
    """Atalho dos três passos, para quando o cabeçalho já é reconhecido."""
    grade = ler_grade(nome_arquivo, conteudo)
    mapa = sugerir_mapa(grade)
    faltando = campos_faltando(mapa)
    if faltando:
        nomes = ", ".join(ROTULOS[c] for c in faltando)
        raise ValueError(f"A planilha precisa ter as colunas Nome, E-mail e Área. Não encontrei: {nomes}.")
    return mapear_e_validar(grade, mapa)
